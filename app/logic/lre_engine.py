#!/usr/bin/env python3
"""
Generador de Libro de Remuneraciones ElectrÃ³nico (LRE)
Fab. de Prod. Metalicos Alejandra Fortuzzi EIRL â RUT: 76.623.639-1

REGLAS APLICADAS (aprendidas en revisiÃ³n con DT):
  1. Formato fecha: dd/mm/aaaa
  2. Separador CSV: punto y coma (;)
  3. CodificaciÃ³n: ANSI / latin-1
  4. Nombre archivo: rutempleadorsinpuntos_AAAAMM.csv
  5. AFP cÃ³digos oficiales: Provida=6, Planvital=11, Capital=31, Modelo=103
  6. Salud: Fonasa=102
  7. Mutual/Org16744: Sin mutual=0, ACHS=1, Mutual CCHC=2, IST=3
  8. RegiÃ³n Metropolitana=13, El Bosque=13105
  9. Jornada Ordinaria Art.22 = 101
 10. Impuesto 2da categorÃ­a = 1
 11. Tramo AF: D=sin derecho, A/B/C segÃºn cargas. REQUERIDO (no puede ir vacÃ­o)
 12. Jubilado (pensionado vejez): SIS=0 en Previred â campo 1109=1, AFC trab=0
 13. AFC trabajador 0.6%: se toma de la liquidaciÃ³n. Si no aparece â 0
     ExcepciÃ³n: jubilados siempre 0 (exentos por ley)
 14. Cotizaciones AFP y Salud: se usan montos del Previred (son los realmente pagados)
 15. Haberes (sueldo, grat, colac, movil, cargas): se suman de todas las hojas del trabajador
 16. Trabajador con mÃºltiples obras â una sola fila LRE con todo sumado
 17. LÃ­quido LRE = Total haberes - Total descuentos (fÃ³rmula DT, puede diferir de liq)
 18. 5301 = suma de todos los 31xx excepto 3164
 19. 5361 = 3161 + 3165
 20. 5341 = 3141+3143+3144+3146+3151+3154+3155+3156+3157+3158
 21. 5302 = 5301 - 5361 - 5362 - 5341
 22. Tasa indemnizaciÃ³n todo evento (1132): vacÃ­o si no aplica
 23. AFC aporte EMPLEADOR: si el "Tipo Contrato" del Previred contiene la
     palabra "Indefinido" â se calcula 2,4% de la renta imponible (sueldo).
     En cualquier otro caso (Plazo Fijo, Por Obra, etc.) se usa el 3% que
     reporta la columna "Seguro CesantÃ­a (AFC)" del Previred, como antes.
"""

import pandas as pd
import openpyxl
import os, re
from datetime import datetime

# ââ CÃDIGOS OFICIALES MANUAL DT âââââââââââââââââââââââââââââââââââââââââââââ
AFP_CODIGOS = {
    'AFP Modelo': 103, 'Modelo': 103,
    'AFP Capital': 31, 'Capital': 31,
    'AFP Habitat': 14, 'Habitat': 14,
    'AFP Planvital': 11, 'Plan Vital': 11, 'Planvital': 11,
    'AFP Provida': 6, 'Provida': 6,
    'AFP Cuprum': 13, 'Cuprum': 13,
    'AFP Uno': 19, 'Uno': 19,
    'No Esta En Afp': 100,
}
SALUD_CODIGOS = {
    'Fonasa': 102, 'FONASA': 102,
    'Cruz Blanca': 1, 'Banmedica': 3, 'Colmena': 4,
    'Consalud': 9, 'Vida Tres': 12,
    'Esencial': 44, 'Nueva Mas Vida': 43,
    'Sin Isapre': 99,
}
MUTUAL_CODIGOS = {
    'Sin Mutual': 0, 'ISL': 0, 'Sin Mutual/Instituto De Seguridad Laboral': 0,
    'ACHS': 1, 'AsociaciÃ³n Chilena de Seguridad': 1,
    'Mutual CCHC': 2, 'Mutual de Seguridad CCHC': 2, 'Mutual CChC': 2,
    'IST': 3, 'Instituto de Seguridad del Trabajo': 3,
}
EMPRESA = {
    'rut_archivo': '766236391',
    'nombre': 'Fab. de Prod. Metalicos Alejandra Fortuzzi EIRL',
    'region': 13,
    'comuna': 13105,
}
COLOR_AZUL  = '#003087'
COLOR_ROJO  = '#C8102E'
COLOR_GRIS  = '#F2F4F7'
COLOR_BLANCO= '#FFFFFF'
COLOR_TEXTO = '#1A1A2E'
COLOR_OK    = '#2E7D32'


# ââ LECTURA DE PREVIRED ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def leer_previred(ruta):
    """
    Lee el consolidado Previred.
    Devuelve dict keyed por RUT normalizado.
    Columnas esperadas: RUT, InstituciÃ³n AFP, InstituciÃ³n Salud, CotizaciÃ³n AFP,
    SIS, CotizaciÃ³n Salud (Fonasa), Seguro CesantÃ­a (AFC), Mutual, DÃ­as Trabajados
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active

    # Encontrar fila de encabezado
    header_row = None
    headers = {}
    for r in ws.iter_rows():
        for cell in r:
            if str(cell.value).strip().upper() == 'RUT':
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        raise ValueError("No se encontrÃ³ encabezado 'RUT' en el Previred.")

    # Mapear columnas
    for cell in ws[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1

    datos = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[headers.get('RUT', 0)]:
            continue
        rut_raw = str(row[headers.get('RUT', 0)]).strip()
        rut = normalizar_rut(rut_raw)
        if not rut:
            continue

        def get(col_name, default=0):
            idx = headers.get(col_name)
            if idx is None:
                return default
            v = row[idx]
            if v is None:
                return default
            try:
                return int(float(str(v).replace(',', '.')))
            except:
                return str(v).strip() if default == '' else default

        # NOTA (corregido): SIS=0 en Previred NO implica automaticamente que
        # el trabajador sea pensionado por vejez (caso real detectado: RUT
        # 7955960-1 tiene SIS=0 y NO es jubilado). Ya no se infiere solo.
        # El estado se deja en 0 por defecto y se marca manualmente en el
        # Paso 3 (checkbox "Pensionado por vejez") para quien corresponda.
        sis = get('SIS', 0)
        pensionado = 0

        # La columna "Seguro CesantÃ­a (AFC)" en Previred es el APORTE EMPLEADOR 3%
        # NO es el descuento del trabajador
        afc_empleador = get('Seguro CesantÃ­a (AFC)', 0)
        mutual_emp    = get('Mutual', 0)

        datos[rut] = {
            'rut': rut,
            'nombre': get('Nombre Completo', ''),
            'afp_nombre': get('InstituciÃ³n AFP', ''),
            'salud_nombre': get('InstituciÃ³n Salud', ''),
            'tipo_contrato': get('Tipo Contrato', ''),
            'dias_previred': get('DÃ­as Trabajados', 30),
            'renta_imponible': get('Renta Imponible', 0),
            'afp_desc': get('CotizaciÃ³n AFP', 0),
            'sis': sis,
            'pensionado': pensionado,
            'salud_desc': get('CotizaciÃ³n Salud (Fonasa)', 0),
            'seguro_social': get('Seguro Social', 0),
            'afc_empleador': afc_empleador,
            'mutual_emp': mutual_emp,
            'cargas_monto': get('Cargas Familiares', 0),
        }

    return datos


# ââ LECTURA DE LIQUIDACIONES âââââââââââââââââââââââââââââââââââââââââââââââââ
def leer_liquidaciones(ruta):
    """
    Lee el libro de liquidaciones (una hoja por trabajador/obra).
    Agrupa por RUT del TRABAJADOR â suma haberes de todas sus hojas.
    El monto siempre estÃ¡ en columna Ã­ndice 6 (col G) de la liquidaciÃ³n.
    """
    COLUMNA_MONTO = 6   # Columna G donde estÃ¡n todos los montos

    wb = openpyxl.load_workbook(ruta, data_only=True)
    resultado = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        filas = list(ws.iter_rows(values_only=True))

        # ââ Extraer RUT del TRABAJADOR ââ
        # Diferencia: fila empresa tiene "AÃ±o" en misma fila, fila trabajador tiene "Seccion"
        rut = None
        fecha_inicio = ''
        nombre_trab = ''

        for fila in filas:
            textos_fila = [str(c).strip() for c in fila if c is not None]
            tiene_anio  = any(t in ('AÃ±o', 'AÃO') for t in textos_fila)
            
            for j, celda in enumerate(fila):
                val = str(celda).strip() if celda else ''
                # RUT del trabajador: etiqueta "Rut" (sin espacios extra) y NO hay "AÃ±o" en la fila
                if val.lower() == 'rut' and not tiene_anio:
                    for k in range(j+1, len(fila)):
                        if fila[k]:
                            rut = normalizar_rut(str(fila[k]))
                            break
                # Fecha de ingreso
                if 'fecha de ingreso' in val.lower() or 'fecha ingreso' in val.lower():
                    for k in range(j+1, len(fila)):
                        if fila[k]:
                            fecha_inicio = parsear_fecha(str(fila[k]))
                            break
                # Nombre del trabajador
                if val.lower() == 'trabajador':
                    for k in range(j+1, len(fila)):
                        if fila[k]:
                            nombre_trab = str(fila[k]).strip()
                            break

        if not rut:
            continue

        # ââ Extraer montos (todos en columna Ã­ndice 6) ââ
        montos = {
            'sueldo': 0, 'gratificacion': 0, 'colacion': 0,
            'movilizacion': 0, 'asig_familiar': 0,
            'afc_trab': 0, 'impuesto': 0, 'dias': 0, 'afp_liq': 0,
        }

        MAPA = {
            'sueldo':        ['SUELDO BASE'],
            'gratificacion': ['GRATIFICACION', 'GRATIFICACIÃN'],
            'colacion':      ['COLACION', 'COLACIÃN'],
            'movilizacion':  ['MOVILIZACION', 'MOVILIZACIÃN'],
            'asig_familiar': ['CARGA FAMILIARES', 'CARGAS FAMILIARES', 'ASIGNACION FAMILIAR'],
            'afc_trab':      ['SEGURO DE CESANTIA', 'SEGURO CESANTIA'],
            'impuesto':      ['IMPUESTO UNICO', 'IMPUESTO ÃNICO', 'IMPUESTO 2DA'],
            'dias':          ['TRABAJADOS'],
            # CORREGIDO: la fila de descuento "AFP <instituciÃ³n>" trae el
            # monto real cotizado en esa liquidaciÃ³n/obra. Se confirmÃ³
            # comparando contra el LRE corregido real que este valor (sumado
            # entre todas las obras del trabajador) es mÃ¡s preciso que la
            # columna "CotizaciÃ³n AFP" del Previred, que a veces trae un
            # monto ligeramente distinto.
            'afp_liq':       ['AFP '],
        }

        for fila in filas:
            # Etiqueta en columna 1, monto en columna 6
            etiq = str(fila[1]).strip().upper() if fila[1] else ''
            if not etiq:
                continue

            for campo, claves in MAPA.items():
                if any(clave in etiq for clave in claves):
                    if campo == 'dias':
                        # DÃ­as estÃ¡ en columna 2 (no en 6)
                        for k in [2, 6, 3, 4, 5]:
                            if k < len(fila) and fila[k] is not None:
                                try:
                                    v = float(str(fila[k]).replace(',','.'))
                                    if 1 <= v <= 31:
                                        montos['dias'] = int(v)
                                        break
                                except:
                                    pass
                    else:
                        # Monto en columna 6
                        if COLUMNA_MONTO < len(fila) and fila[COLUMNA_MONTO] is not None:
                            try:
                                v = float(str(fila[COLUMNA_MONTO]).replace(',','.'))
                                if v > 0:
                                    montos[campo] = int(round(v))
                            except:
                                pass
                    break

        # ââ Acumular en resultado ââ
        if rut not in resultado:
            resultado[rut] = {
                'rut': rut,
                'nombre': nombre_trab,
                'fecha_inicio': fecha_inicio,
                'sueldo': 0, 'gratificacion': 0, 'colacion': 0,
                'movilizacion': 0, 'asig_familiar': 0,
                'afc_trab': 0, 'impuesto': 0, 'dias': 0, 'hojas': 0,
                'afp_liq': 0,
            }

        resultado[rut]['sueldo']        += montos['sueldo']
        resultado[rut]['gratificacion'] += montos['gratificacion']
        resultado[rut]['colacion']      += montos['colacion']
        resultado[rut]['movilizacion']  += montos['movilizacion']
        resultado[rut]['asig_familiar'] += montos['asig_familiar']
        resultado[rut]['afc_trab']      += montos['afc_trab']
        resultado[rut]['impuesto']      += montos['impuesto']
        resultado[rut]['dias']          += montos['dias']
        resultado[rut]['afp_liq']       += montos['afp_liq']
        resultado[rut]['hojas']         += 1

        if not resultado[rut]['fecha_inicio'] and fecha_inicio:
            resultado[rut]['fecha_inicio'] = fecha_inicio
        if not resultado[rut]['nombre'] and nombre_trab:
            resultado[rut]['nombre'] = nombre_trab

    # Limitar dÃ­as a 30 (regla DT)
    for rut in resultado:
        resultado[rut]['dias'] = min(resultado[rut]['dias'], 30)

    return resultado


def extraer_montos_liquidacion(filas):
    """
    Busca filas clave en la liquidaciÃ³n y extrae montos.
    El monto estÃ¡ casi siempre en la columna 6 (Ã­ndice 6) de la fila.
    """
    res = {
        'sueldo': 0, 'gratificacion': 0, 'colacion': 0,
        'movilizacion': 0, 'asig_familiar': 0,
        'afc_trab': 0, 'impuesto': 0, 'dias': 0,
    }

    CLAVES = {
        'sueldo':        ['SUELDO BASE', 'SUELDO'],
        'gratificacion': ['GRATIFICACION', 'GRATIFICACIÃN'],
        'colacion':      ['COLACION', 'COLACIÃN'],
        'movilizacion':  ['MOVILIZACION', 'MOVILIZACIÃN'],
        'asig_familiar': ['CARGA FAMILIARES', 'CARGAS FAMILIARES', 'ASIGNACION FAMILIAR'],
        'afc_trab':      ['SEGURO DE CESANTIA', 'SEGURO CESANTIA', 'AFC TRABAJADOR'],
        'impuesto':      ['IMPUESTO UNICO', 'IMPUESTO ÃNICO', 'IMPUESTO 2DA', 'IMP. UNICO'],
        'dias':          ['TRABAJADOS'],
    }

    for fila in filas:
        texto_fila = ' | '.join(str(c).upper().strip() for c in fila if c)
        for campo, claves in CLAVES.items():
            for clave in claves:
                if clave in texto_fila:
                    # Buscar el primer nÃºmero en la fila (de derecha a izquierda para montos)
                    monto = extraer_numero_fila(fila, campo == 'dias')
                    if monto and monto > 0:
                        if campo == 'dias':
                            res['dias'] = int(monto)
                        else:
                            res[campo] = int(round(monto))
                    break

    return res


def extraer_numero_fila(fila, es_dias=False):
    """Extrae el primer nÃºmero significativo de una fila."""
    candidatos = []
    for celda in fila:
        if celda is None:
            continue
        try:
            v = float(str(celda).replace(',', '.'))
            if es_dias:
                if 1 <= v <= 31:
                    candidatos.append(v)
            else:
                if v > 100:  # montos significativos
                    candidatos.append(v)
        except:
            pass
    return candidatos[0] if candidatos else None


def normalizar_rut(rut_raw):
    """Normaliza un RUT a formato 12345678-9 sin puntos."""
    if not rut_raw:
        return None
    rut = str(rut_raw).strip().upper()
    rut = rut.replace('.', '').replace(' ', '')
    # Si tiene guion â ya estÃ¡ bien formateado
    if '-' in rut:
        partes = rut.split('-')
        if len(partes) == 2:
            num = partes[0].lstrip('0')
            dv  = partes[1]
            return f'{num}-{dv}' if num else None
    # Sin guion: Ãºltimo char es DV
    if len(rut) >= 2:
        return f'{rut[:-1].lstrip("0")}-{rut[-1]}'
    return None


def parsear_fecha(texto):
    """
    Convierte texto de fecha a formato dd/mm/aaaa.
    Acepta: '01/04/2025', '1 de Abril de 2025', '20250401', datetime objects.
    """
    if not texto or str(texto).strip() in ('', 'None'):
        return ''

    texto = str(texto).strip()

    # Si ya es datetime
    try:
        if hasattr(texto, 'strftime'):
            return texto.strftime('%d/%m/%Y')
    except:
        pass

    # Formato dd/mm/aaaa o dd-mm-aaaa
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', texto)
    if m:
        return f'{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}'

    # Formato aaaammdd
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', texto)
    if m:
        return f'{m.group(3)}/{m.group(2)}/{m.group(1)}'

    # Texto tipo "01 de Abril de 2025" o "01 de Enero 2026"
    meses_es = {
        'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
        'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12
    }
    m = re.search(r'(\d{1,2})\s+de\s+(\w+)\s+(?:de\s+)?(\d{4})', texto.lower())
    if m:
        dia = int(m.group(1))
        mes = meses_es.get(m.group(2), 0)
        anio = int(m.group(3))
        if mes:
            return f'{dia:02d}/{mes:02d}/{anio}'

    # Intentar parsear con pandas
    try:
        dt = pd.to_datetime(texto, dayfirst=True)
        return dt.strftime('%d/%m/%Y')
    except:
        pass

    return texto  # devolver como estÃ¡ si no se pudo parsear


# ââ CONSOLIDAR DATOS âââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def consolidar(previred, liquidaciones):
    """
    Une datos de Previred + liquidaciones para cada trabajador.
    Aplica todas las reglas de negocio aprendidas.
    """
    trabajadores = []

    for rut, prev in previred.items():
        liq = liquidaciones.get(rut, {})

        # ââ Haberes ââ
        # CORREGIDO: El campo Sueldo(2101) de la LRE debe ir SIEMPRE con la
        # Renta Imponible del Previred (que ya incluye sueldo + gratificaciÃ³n
        # combinados). Se comprobÃ³, comparando contra el LRE corregido real,
        # que la DT/Previred no separa sueldo y gratificaciÃ³n en el LRE: todo
        # el monto imponible va en 2101 y el campo 2106 (GratificaciÃ³n) queda
        # vacÃ­o. Por eso ya NO se usa el desglose sueldo+gratificaciÃ³n de la
        # liquidaciÃ³n para el campo 2101; se usa el imponible del Previred.
        sueldo        = prev['renta_imponible']
        gratificacion = 0
        colacion      = liq.get('colacion', 0)
        movilizacion  = liq.get('movilizacion', 0)
        asig_familiar = liq.get('asig_familiar', 0)

        # ââ AFC del trabajador: SOLO si aparece explÃ­cito en la liquidaciÃ³n ââ
        # CORREGIDO: ya no se calcula al 0.6% cuando no aparece; si la
        # liquidaciÃ³n no trae el monto, se deja en 0 (y se exporta vacÃ­o).
        # Jubilado (pensionado_vejez=1) â siempre 0, exento por ley.
        if prev['pensionado'] == 1:
            afc_trab = 0
        else:
            afc_trab = liq.get('afc_trab', 0)

        # ââ CotizaciÃ³n AFP: preferir suma real de las liquidaciones ââ
        # CORREGIDO: si la(s) liquidaciÃ³n(es) traen el monto de AFP, se usa
        # esa suma (mÃ¡s precisa, prorateada por obra/dÃ­as). Si no aparece,
        # se usa como respaldo la "CotizaciÃ³n AFP" del Previred.
        afp_desc  = liq.get('afp_liq', 0) or prev['afp_desc']
        salud_desc= prev['salud_desc']
        impuesto  = liq.get('impuesto', 0)

        # ââ DÃ­as trabajados ââ
        # El manual dice: trabajador mensual que trabajÃ³ todos los dÃ­as â 30
        dias = liq.get('dias', 0) or prev.get('dias_previred', 30)
        # Limitar a 30 mÃ¡ximo (regla DT)
        dias = min(int(dias), 30)

        # ââ Cargas familiares ââ
        cargas_legales = 0
        tramo_af = 'D'
        if asig_familiar > 0 or prev.get('cargas_monto', 0) > 0:
            cargas_legales = 1  # al menos 1 carga
            tramo_af = 'A'      # asumir tramo A (el mÃ¡s comÃºn)
            if asig_familiar == 0:
                asig_familiar = prev.get('cargas_monto', 0)

        # ââ Mutual: del Previred ââ
        mutual_nombre = inferir_mutual(prev.get('salud_nombre', ''),
                                       prev.get('afp_nombre', ''))
        mutual_cod = MUTUAL_CODIGOS.get(mutual_nombre, 2)

        # ââ AFP y Salud: cÃ³digos oficiales ââ
        afp_cod   = AFP_CODIGOS.get(prev['afp_nombre'], 6)
        salud_cod = SALUD_CODIGOS.get(prev['salud_nombre'], 102)
        if 'fonasa' in prev['salud_nombre'].lower():
            salud_cod = 102

        # ââ Aportes empleador ââ
        # CORREGIDO (confirmado con datos reales de mayo Y junio, ambos
        # meses coinciden): cuando el Previred reporta SIS=0 para un
        # trabajador, la ley exime al empleador tanto del aporte SIS(4155)
        # como del aporte AFC(4151) para ese trabajador â es el caso de un
        # trabajador pensionado que reingresÃ³ a trabajar, aunque el
        # checkbox "Pensionado por vejez(1109)" de la LRE se mantenga en 0.
        # Esta exenciÃ³n es automÃ¡tica segÃºn el propio Previred (SIS=0) y
        # NO depende del checkbox manual de pensionado.
        # ââ Aporte AFC empleador: depende del tipo de contrato ââ
        # Contrato INDEFINIDO: empleador aporta 2,4% de la renta imponible
        # (1,6% cuenta individual + 0,8% fondo solidario). Trabajador aporta
        # 0,6% (afc_trab, ya calculado mÃ¡s arriba desde la liquidaciÃ³n).
        # Contrato PLAZO FIJO / POR OBRA: empleador aporta 3% (columna
        # "Seguro CesantÃ­a (AFC)" del Previred), trabajador no aporta.
        es_indefinido = 'indefinid' in str(prev.get('tipo_contrato', '')).lower()

        sis = prev.get('sis', 0)
        if sis == 0:
            sis_emp = 0
            afc_emp = 0
        else:
            sis_emp = sis
            if es_indefinido:
                afc_emp = round(sueldo * 0.024)
            else:
                afc_emp = prev.get('afc_empleador', 0)
        mutual_emp = prev.get('mutual_emp', 0)
        # CORREGIDO: la columna "Seguro Social" del Previred corresponde al
        # aporte de indemnizaciÃ³n a todo evento del empleador (campo 4131).
        indemnizacion_emp = prev.get('seguro_social', 0)
        # Si ademÃ¡s el usuario marca manualmente "Pensionado por vejez",
        # tambiÃ©n se exime el aporte SIS (por si el Previred no lo reflejara)
        if prev['pensionado'] == 1:
            sis_emp = 0

        trabajadores.append({
            'rut': rut,
            'nombre': prev['nombre'],
            'fecha_inicio': liq.get('fecha_inicio', ''),
            'fecha_termino': '',
            'causal_termino': '',
            'afp_cod': afp_cod,
            'afp_nombre': prev['afp_nombre'],
            'salud_cod': salud_cod,
            'salud_nombre': prev['salud_nombre'],
            'mutual_cod': mutual_cod,
            'mutual_nombre': mutual_nombre,
            'pensionado': prev['pensionado'],
            'dias': dias,
            'tramo_af': tramo_af,
            'cargas_legales': cargas_legales,
            'sueldo': sueldo,
            'gratificacion': gratificacion,
            'colacion': colacion,
            'movilizacion': movilizacion,
            'asig_familiar': asig_familiar,
            'afp_desc': afp_desc,
            'salud_desc': salud_desc,
            'afc_trab': afc_trab,
            'impuesto': impuesto,
            'afc_emp': afc_emp,
            'mutual_emp': mutual_emp,
            'sis_emp': sis_emp,
            'indemnizacion_emp': indemnizacion_emp,
        })

    return trabajadores


def inferir_mutual(salud_nombre, afp_nombre):
    """Infiere la mutual segÃºn la instituciÃ³n de salud."""
    s = salud_nombre.upper()
    if 'MUTUAL' in s or 'CCHC' in s:
        return 'Mutual CCHC'
    if 'ACHS' in s:
        return 'ACHS'
    if 'IST' in s:
        return 'IST'
    # Fonasa â normalmente sin mutual (ISL)
    # PERO en este empresa todos usan Mutual CCHC independiente de salud
    # El Previred trae la columna "Mutual" con el monto â si tiene monto â Mutual CCHC
    return 'Mutual CCHC'  # default para esta empresa


# ââ CONSTRUIR FILA LRE âââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def construir_fila_lre(t):
    """Construye el diccionario completo de 147 campos segÃºn manual DT."""
    imp_trib       = int(t['sueldo']) + int(t['gratificacion'])
    no_imp_no_trib = int(t['colacion']) + int(t['movilizacion']) + int(t['asig_familiar'])
    total_haberes  = imp_trib + no_imp_no_trib

    total_cotiz = (int(t['afp_desc']) + int(t['salud_desc']) + int(t['afc_trab']))
    total_imp_r = int(t['impuesto'])
    # 5301 = suma todos 31xx excepto 3164
    total_desc  = total_cotiz + total_imp_r
    # 5302 = 5301 - 5361 - 5362 - 5341
    otros_desc  = 0

    total_aportes = (int(t['afc_emp']) + int(t['mutual_emp']) + int(t['sis_emp'])
                      + int(t.get('indemnizacion_emp', 0)))
    liquido = total_haberes - total_desc

    fila = {
        'Rut trabajador(1101)': t['rut'],
        'Fecha inicio contrato(1102)': t.get('fecha_inicio', ''),
        'Fecha tÃ©rmino de contrato(1103)': t.get('fecha_termino', ''),
        'Causal tÃ©rmino de contrato(1104)': t.get('causal_termino', ''),
        'RegiÃ³n prestaciÃ³n de servicios(1105)': EMPRESA['region'],
        'Comuna prestaciÃ³n de servicios(1106)': EMPRESA['comuna'],
        'Tipo impuesto a la renta(1170)': 1,
        'TÃ©cnico extranjero exenciÃ³n cot. previsionales(1146)': 0,
        'CÃ³digo tipo de jornada(1107)': 101,
        'Persona con Discapacidad - Pensionado por Invalidez(1108)': 0,
        'Pensionado por vejez(1109)': int(t['pensionado']),
        'AFP(1141)': int(t['afp_cod']),
        'IPS (ExINP)(1142)': 0,
        'FONASA - ISAPRE(1143)': int(t['salud_cod']),
        'AFC(1151)': 1,
        'CCAF(1110)': 0,
        'Org. administrador ley 16.744(1152)': int(t['mutual_cod']),
        'Nro cargas familiares legales autorizadas(1111)': (int(t['cargas_legales']) if t['cargas_legales'] else ''),
        'Nro de cargas familiares maternales(1112)': 0,
        'Nro de cargas familiares invalidez(1113)': 0,
        'Tramo asignaciÃ³n familiar(1114)': (t['tramo_af'] if t['cargas_legales'] else ''),
        'Rut org sindical 1(1171)': '', 'Rut org sindical 2(1172)': '',
        'Rut org sindical 3(1173)': '', 'Rut org sindical 4(1174)': '',
        'Rut org sindical 5(1175)': '', 'Rut org sindical 6(1176)': '',
        'Rut org sindical 7(1177)': '', 'Rut org sindical 8(1178)': '',
        'Rut org sindical 9(1179)': '', 'Rut org sindical 10(1180)': '',
        'Nro dÃ­as trabajados en el mes(1115)': int(t['dias']),
        'Nro dÃ­as de licencia mÃ©dica en el mes(1116)': 0,
        'Nro dÃ­as de vacaciones en el mes(1117)': 0,
        'Subsidio trabajador joven(1118)': 0,
        'Puesto Trabajo Pesado(1154)': 0,
        'APVI(1155)': 0,
        'APVC(1157)': 0,
        'IndemnizaciÃ³n a todo evento(1131)': 0,
        'Tasa indemnizaciÃ³n a todo evento(1132)': '',
        'Sueldo(2101)': int(t['sueldo']),
        'Sobresueldo(2102)': 0, 'Comisiones(2103)': 0,
        'Semana corrida(2104)': 0, 'ParticipaciÃ³n(2105)': 0,
        'GratificaciÃ³n(2106)': int(t['gratificacion']),
        'Recargo 30% dÃ­a domingo(2107)': 0,
        'Remun. variable pagada en vacaciones(2108)': 0,
        'Remun. variable pagada en clausura(2109)': 0,
        'Aguinaldo(2110)': 0,
        'Bonos u otras remun. fijas mensuales(2111)': 0,
        'Tratos(2112)': 0,
        'Bonos u otras remun. variables mensuales o superiores a un mes(2113)': 0,
        'Ejercicio opciÃ³n no pactada en contrato(2114)': 0,
        'Beneficios en especie constitutivos de remun(2115)': 0,
        'Remuneraciones bimestrales(2116)': 0, 'Remuneraciones trimestrales(2117)': 0,
        'Remuneraciones cuatrimestral(2118)': 0, 'Remuneraciones semestrales(2119)': 0,
        'Remuneraciones anuales(2120)': 0, 'ParticipaciÃ³n anual(2121)': 0,
        'GratificaciÃ³n anual(2122)': 0,
        'Otras remuneraciones superiores a un mes(2123)': 0,
        'Pago por horas de trabajo sindical(2124)': 0,
        'Sueldo empresarial (2161)': 0,
        'Subsidio por incapacidad laboral por licencia mÃ©dica(2201)': 0,
        'Beca de estudio(2202)': 0, 'Gratificaciones de zona(2203)': 0,
        'Otros ingresos no constitutivos de renta(2204)': 0,
        'ColaciÃ³n(2301)': int(t['colacion']),
        'MovilizaciÃ³n(2302)': int(t['movilizacion']),
        'ViÃ¡ticos(2303)': 0, 'AsignaciÃ³n de pÃ©rdida de caja(2304)': 0,
        'AsignaciÃ³n de desgaste herramienta(2305)': 0,
        'AsignaciÃ³n familiar legal(2311)': int(t['asig_familiar']),
        'Gastos por causa del trabajo(2306)': 0,
        'Gastos por cambio de residencia(2307)': 0,
        'Sala cuna(2308)': 0,
        'AsignaciÃ³n trabajo a distancia o teletrabajo(2309)': 0,
        'DepÃ³sito convenido hasta UF 900(2347)': 0,
        'Alojamiento por razones de trabajo(2310)': 0,
        'AsignaciÃ³n de traslaciÃ³n(2312)': 0,
        'IndemnizaciÃ³n por feriado legal(2313)': 0,
        'IndemnizaciÃ³n aÃ±os de servicio(2314)': 0,
        'IndemnizaciÃ³n sustitutiva del aviso previo(2315)': 0,
        'IndemnizaciÃ³n fuero maternal(2316)': 0,
        'Pago indemnizaciÃ³n a todo evento(2331)': 0,
        'Indemnizaciones voluntarias tributables(2417)': 0,
        'Indemnizaciones contractuales tributables(2418)': 0,
        'CotizaciÃ³n obligatoria previsional (AFP o IPS)(3141)': int(t['afp_desc']),
        'CotizaciÃ³n obligatoria salud 7%(3143)': int(t['salud_desc']),
        'CotizaciÃ³n voluntaria para salud(3144)': 0,
        'CotizaciÃ³n AFC - trabajador(3151)': int(t['afc_trab']),
        'Cotizaciones tÃ©cnico extranjero para seguridad social fuera de Chile(3146)': 0,
        'Descuento depÃ³sito convenido hasta UF 900 anual(3147)': 0,
        'CotizaciÃ³n APVi Mod A(3155)': 0, 'CotizaciÃ³n APVi Mod B hasta UF50(3156)': 0,
        'CotizaciÃ³n APVc Mod A(3157)': 0, 'CotizaciÃ³n APVc Mod B hasta UF50(3158)': 0,
        'Impuesto retenido por remuneraciones(3161)': int(t['impuesto']),
        'Impuesto retenido por indemnizaciones(3162)': 0,
        'Mayor retenciÃ³n de impuestos solicitada por el trabajador(3163)': 0,
        'Impuesto retenido por reliquidaciÃ³n remun. devengadas otros perÃ­odos(3164)': 0,
        'Diferencia impuesto reliquidaciÃ³n remun. devengadas en este perÃ­odo(3165)': 0,
        'RetenciÃ³n prÃ©stamo clase media 2020 (Ley 21.252) (3166)': 0,
        'Rebaja zona extrema DL 889 (3167)': 0,
        'Cuota sindical 1(3171)': 0, 'Cuota sindical 2(3172)': 0,
        'Cuota sindical 3(3173)': 0, 'Cuota sindical 4(3174)': 0,
        'Cuota sindical 5(3175)': 0, 'Cuota sindical 6(3176)': 0,
        'Cuota sindical 7(3177)': 0, 'Cuota sindical 8(3178)': 0,
        'Cuota sindical 9(3179)': 0, 'Cuota sindical 10(3180)': 0,
        'CrÃ©dito social CCAF(3110)': 0, 'Cuota vivienda o educaciÃ³n(3181)': 0,
        'CrÃ©dito cooperativas de ahorro(3182)': 0,
        'Otros descuentos autorizados y solicitados por el trabajador(3183)': 0,
        'CotizaciÃ³n adicional trabajo pesado - trabajador(3154)': 0,
        'Donaciones culturales y de reconstrucciÃ³n(3184)': 0,
        'Otros descuentos(3185)': 0, 'Pensiones de alimentos(3186)': 0,
        'Descuento mujer casada(3187)': 0,
        'Descuentos por anticipos y prÃ©stamos(3188)': 0,
        'AFC - Aporte empleador(4151)': int(t['afc_emp']),
        'Aporte empleador seguro accidentes del trabajo y Ley SANNA(4152)': int(t['mutual_emp']),
        'Aporte empleador indemnizaciÃ³n a todo evento(4131)': int(t.get('indemnizacion_emp', 0)),
        'Aporte adicional trabajo pesado - empleador(4154)': 0,
        'Aporte empleador seguro invalidez y sobrevivencia(4155)': int(t['sis_emp']),
        'APVC - Aporte Empleador(4157)': 0,
        'Total haberes(5201)': total_haberes,
        'Total haberes imponibles y tributables(5210)': imp_trib,
        'Total haberes imponibles no tributables(5220)': 0,
        'Total haberes no imponibles y no tributables(5230)': no_imp_no_trib,
        'Total haberes no imponibles y tributables(5240)': 0,
        'Total descuentos(5301)': total_desc,
        'Total descuentos impuestos a las remuneraciones(5361)': total_imp_r,
        'Total descuentos impuestos por indemnizaciones(5362)': 0,
        'Total descuentos por cotizaciones del trabajador(5341)': total_cotiz,
        'Total otros descuentos(5302)': otros_desc,
        'Total aportes empleador(5410)': total_aportes,
        'Total lÃ­quido(5501)': liquido,
        'Total indemnizaciones(5502)': 0,
        'Total indemnizaciones tributables(5564)': 0,
        'Total indemnizaciones no tributables(5565)': 0,
    }

    # ââ CORREGIDO: dejar en BLANCO los campos que la empresa nunca usa ââââ
    # Al comparar contra el LRE corregido real de la DT, se confirmÃ³ que
    # todos estos campos (haberes/descuentos/aportes/contadores que esta
    # empresa jamÃ¡s utiliza) deben quedar con la celda VACÃA, no con "0".
    # A diferencia de un enfoque "blanquear todo lo que sea 0", aquÃ­ se usa
    # una lista explÃ­cita para NO tocar campos que sÃ­ son de uso real de la
    # empresa (sueldo, colaciÃ³n, cotizaciones, aportes, totales, etc.) aun
    # cuando su valor puntual sea 0 en algÃºn trabajador â esos deben seguir
    # mostrando "0" y no quedar vacÃ­os.
    campos_blanquear_si_cero = {
        # Haberes que esta empresa nunca paga
        'Sobresueldo(2102)', 'Comisiones(2103)', 'Semana corrida(2104)',
        'ParticipaciÃ³n(2105)', 'GratificaciÃ³n(2106)',
        'Recargo 30% dÃ­a domingo(2107)',
        'Remun. variable pagada en vacaciones(2108)',
        'Remun. variable pagada en clausura(2109)', 'Aguinaldo(2110)',
        'Bonos u otras remun. fijas mensuales(2111)', 'Tratos(2112)',
        'Bonos u otras remun. variables mensuales o superiores a un mes(2113)',
        'Ejercicio opciÃ³n no pactada en contrato(2114)',
        'Beneficios en especie constitutivos de remun(2115)',
        'Remuneraciones bimestrales(2116)', 'Remuneraciones trimestrales(2117)',
        'Remuneraciones cuatrimestral(2118)', 'Remuneraciones semestrales(2119)',
        'Remuneraciones anuales(2120)', 'ParticipaciÃ³n anual(2121)',
        'GratificaciÃ³n anual(2122)',
        'Otras remuneraciones superiores a un mes(2123)',
        'Pago por horas de trabajo sindical(2124)', 'Sueldo empresarial (2161)',
        'Subsidio por incapacidad laboral por licencia mÃ©dica(2201)',
        'Beca de estudio(2202)', 'Gratificaciones de zona(2203)',
        'Otros ingresos no constitutivos de renta(2204)',
        'ViÃ¡ticos(2303)', 'AsignaciÃ³n de pÃ©rdida de caja(2304)',
        'AsignaciÃ³n de desgaste herramienta(2305)',
        'Gastos por causa del trabajo(2306)',
        'Gastos por cambio de residencia(2307)', 'Sala cuna(2308)',
        'AsignaciÃ³n trabajo a distancia o teletrabajo(2309)',
        'Alojamiento por razones de trabajo(2310)',
        'AsignaciÃ³n de traslaciÃ³n(2312)',
        'IndemnizaciÃ³n por feriado legal(2313)',
        'IndemnizaciÃ³n aÃ±os de servicio(2314)',
        'IndemnizaciÃ³n sustitutiva del aviso previo(2315)',
        'IndemnizaciÃ³n fuero maternal(2316)',
        'DepÃ³sito convenido hasta UF 900(2347)',
        'Pago indemnizaciÃ³n a todo evento(2331)',
        'Indemnizaciones voluntarias tributables(2417)',
        'Indemnizaciones contractuales tributables(2418)',
        # Descuentos que esta empresa nunca aplica
        'CotizaciÃ³n voluntaria para salud(3144)',
        'Cotizaciones tÃ©cnico extranjero para seguridad social fuera de Chile(3146)',
        'Descuento depÃ³sito convenido hasta UF 900 anual(3147)',
        'CotizaciÃ³n APVi Mod A(3155)', 'CotizaciÃ³n APVi Mod B hasta UF50(3156)',
        'CotizaciÃ³n APVc Mod A(3157)', 'CotizaciÃ³n APVc Mod B hasta UF50(3158)',
        'Impuesto retenido por indemnizaciones(3162)',
        'Mayor retenciÃ³n de impuestos solicitada por el trabajador(3163)',
        'Impuesto retenido por reliquidaciÃ³n remun. devengadas otros perÃ­odos(3164)',
        'Diferencia impuesto reliquidaciÃ³n remun. devengadas en este perÃ­odo(3165)',
        'RetenciÃ³n prÃ©stamo clase media 2020 (Ley 21.252) (3166)',
        'Rebaja zona extrema DL 889 (3167)',
        'Cuota sindical 1(3171)', 'Cuota sindical 2(3172)',
        'Cuota sindical 3(3173)', 'Cuota sindical 4(3174)',
        'Cuota sindical 5(3175)', 'Cuota sindical 6(3176)',
        'Cuota sindical 7(3177)', 'Cuota sindical 8(3178)',
        'Cuota sindical 9(3179)', 'Cuota sindical 10(3180)',
        'CrÃ©dito social CCAF(3110)', 'Cuota vivienda o educaciÃ³n(3181)',
        'CrÃ©dito cooperativas de ahorro(3182)',
        'Otros descuentos autorizados y solicitados por el trabajador(3183)',
        'CotizaciÃ³n adicional trabajo pesado - trabajador(3154)',
        'Donaciones culturales y de reconstrucciÃ³n(3184)',
        'Otros descuentos(3185)', 'Pensiones de alimentos(3186)',
        'Descuento mujer casada(3187)',
        'Descuentos por anticipos y prÃ©stamos(3188)',
        # Aportes empleador que esta empresa nunca usa
        'Aporte adicional trabajo pesado - empleador(4154)',
        'APVC - Aporte Empleador(4157)',
        # Contadores/indicadores que esta empresa nunca usa
        'Nro de cargas familiares maternales(1112)',
        'Nro de cargas familiares invalidez(1113)',
        'Nro dÃ­as de licencia mÃ©dica en el mes(1116)',
        'Nro dÃ­as de vacaciones en el mes(1117)',
        'Puesto Trabajo Pesado(1154)',
        # AFC trabajador: si no viene explÃ­cito en la liquidaciÃ³n, vacÃ­o
        # (ya NO se calcula al 0.6% automÃ¡ticamente)
        'CotizaciÃ³n AFC - trabajador(3151)',
        'AsignaciÃ³n familiar legal(2311)',
        # Totales de indemnizaciones que esta empresa nunca usa
        'Total indemnizaciones(5502)',
        'Total indemnizaciones no tributables(5565)',
        'Total descuentos impuestos por indemnizaciones(5362)',
    }
    for clave in campos_blanquear_si_cero:
        if clave in fila and fila[clave] == 0:
            fila[clave] = ''

    return fila


# ââ INTERFAZ GRÃFICA âââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

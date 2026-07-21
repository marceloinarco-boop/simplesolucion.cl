"""
=============================================================================
PROCESADOR DE PLANILLAS PREVIRED - Chile  (v4 - Con días trabajados)
=============================================================================
Librerías requeridas:
  pip install pdfplumber pandas openpyxl

Ejecutar:
  python previred_processor.py
=============================================================================

LÓGICA DE CONSOLIDACIÓN:
  Cada PDF contiene el mismo trabajador en 3+ documentos (AFP, Fonasa, Mutual).
  Si sumáramos la renta de todos, se multiplicaría x3 o más.

  Solución: cada fuente aporta SOLO el campo que le pertenece:
    - AFP detalle  → renta_imponible  +  cotizacion_afp  +  afc
    - Fonasa anexo → cotizacion_salud  (la renta ya viene del AFP)
    - Mutual       → cotizacion_mutual (la renta ya viene del AFP)
    - Seguro Social→ seg_social        (informativo, no duplica renta)

  Al consolidar por RUT se suman solo los campos que cada fuente aporta.
  La renta imponible viene SOLO del documento AFP (una sola vez por período).
=============================================================================
"""

import os
import re

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 1: UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def limpiar_rut(rut_raw: str) -> str:
    """Normaliza RUT: quita puntos, guiones, espacios → '12345678K'."""
    return re.sub(r"[\.\-\s]", "", str(rut_raw)).upper().strip()


def es_rut_valido(txt: str) -> bool:
    """Verifica formato básico de RUT chileno."""
    return bool(re.match(r"^\d{1,2}\.?\d{3}\.?\d{3}[-]?[\dkK]$", str(txt).strip()))


def a_numero(valor) -> float:
    """Convierte string con puntos de miles a float. Ej: '1.234.567' → 1234567.0"""
    if valor is None:
        return 0.0
    s = re.sub(r"[^\d]", "", str(valor))
    return float(s) if s else 0.0


def nombre_completo(*partes) -> str:
    """Une apellidos y nombres en string limpio."""
    return " ".join(str(p).strip() for p in partes if p and str(p).strip())


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 2: DETECCIÓN DEL TIPO DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────

def detectar_tipo_pagina(texto: str) -> str:
    """
    Identifica qué tipo de comprobante Previred contiene la página.
    Cada tipo aporta campos distintos al registro final.

    Retorna: 'afp_detalle' | 'mutual' | 'fonasa_anexo' | 'seguro_social' | 'otro'
    """
    t = texto.lower()

    # AFP Detalle: tiene RUT + nombre + Remuneración Imponible + Cotización Obligatoria
    if ("detalle de pago de cotizaciones" in t and
            ("afp" in t or "fondo de pensiones" in t) and
            "remuneración" in t and "cotización" in t):
        return "afp_detalle"

    # Mutual: "detalle de comprobante pago de cotizaciones" + mutual
    if ("detalle de comprobante pago" in t and "mutual" in t):
        return "mutual"

    # IPS Anexo Trabajadores: contiene cargas familiares (asignación familiar)
    if ("tr ax anexo trabajadores" in t or
            ("ips" in t and "asignacion familiar" in t and "apellido paterno" in t)):
        return "ips_anexo"

    # Fonasa Anexo: "anexos de detalle" con tabla de cotizaciones de salud
    if ("anexos de detalle" in t and
            ("fonasa" in t or "cotización 7%" in t or "cotizacion 7%" in t
             or "cotiz" in t)):
        return "fonasa_anexo"

    # Seguro Social detalle — la página empieza con "Pago Electrónico" y luego
    # "Identificación del Trabajador" con "Nombres y Apellidos"
    if ("seguro social" in t and
            "nombres y apellidos" in t and
            "renta imponible" in t):
        return "seguro_social"

    return "otro"


def detectar_afp_en_texto(texto: str) -> str:
    """Extrae el nombre de la AFP del texto de la página."""
    m = re.search(
        r"AFP\s+(Provida|Capital|Modelo|PlanVital|Uno|Cuprum|Habitat|Ciedess)",
        texto, re.IGNORECASE
    )
    if m:
        return "AFP " + m.group(1).strip().title()
    # Fallback: buscar solo el nombre
    m2 = re.search(r"\b(Provida|Capital|Modelo|PlanVital|AFP UNO|Cuprum|Habitat)\b",
                   texto, re.IGNORECASE)
    if m2:
        return "AFP " + m2.group(1).strip().title()
    return "AFP"


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 3: PARSERS POR TIPO — cada uno aporta SOLO sus campos propios
# ─────────────────────────────────────────────────────────────────────────────

# ── Lista de RUT con contrato INDEFINIDO ─────────────────────────────────────
# Todos los RUT que NO estén aquí se consideran contrato PLAZO FIJO.
# Formato: sin puntos, sin guión, en mayúsculas.
# ─────────────────────────────────────────────────────────────────────────────
RUT_INDEFINIDO = {
    "132976252",   # BARRAZA BUSTAMANTE ALEX DAVID
    "139348532",   # FORTUZZI ULLOA ALEJANDRA VALESKA
}

FILAS_IGNORAR = {
    "rut", "apellido", "nombres", "nombre", "totales", "total",
    "identificacion", "identificación", "n°", "n", "remuneración",
    "remuneracion", "", "totales generales", "total página", "total acumulado"
}

def es_encabezado(fila: list) -> bool:
    primera = str(fila[0]).strip().lower() if fila and fila[0] else ""
    return primera in FILAS_IGNORAR


def _rut_offset(fila: list):
    """
    Busca en qué columna está el RUT (puede estar en col 0 o col 1 si hay N° de fila).
    Retorna (rut_raw, offset) o (None, None).
    """
    for i in range(min(2, len(fila))):
        candidato = str(fila[i]).strip() if fila[i] else ""
        if es_rut_valido(candidato):
            return candidato, i
    return None, None


# ── AFP Detalle ──────────────────────────────────────────────────────────────
# Columnas: RUT | Nombres | Rem.Imponible | Cot.Obligatoria | SIS | CotVol |
#           N°Contrato | DepConv | DepCta | Rem.Imp(AFC) | Cot.Afiliado(AFC)
#
# Esta fuente aporta: renta + cot_afp + afc + nombre + inst_afp
# ────────────────────────────────────────────────────────────────────────────

def parsear_afp(tabla: list, nombre_afp: str) -> list[dict]:
    """
    Tabla AFP detalle.
    Tipo contrato se determina por lista fija de RUT (ver RUT_PLAZO_FIJO).
    """
    registros = []
    for fila in tabla:
        if not fila or es_encabezado(fila):
            continue
        rut_raw, off = _rut_offset(fila)
        if not rut_raw:
            continue

        nombre_raw  = str(fila[off + 1]).strip() if len(fila) > off + 1 else ""
        renta       = a_numero(fila[off + 2]) if len(fila) > off + 2 else 0.0
        cot_afp     = a_numero(fila[off + 3]) if len(fila) > off + 3 else 0.0
        sis         = a_numero(fila[off + 4]) if len(fila) > off + 4 else 0.0

        if renta == 0 and cot_afp == 0:
            continue

        rut_limpio = limpiar_rut(rut_raw)
        tipo = "Indefinido" if rut_limpio in RUT_INDEFINIDO else "Plazo Fijo"

        registros.append({
            "fuente":        "afp",
            "rut":           rut_limpio,
            "nombre":        nombre_raw,
            "inst_afp":      nombre_afp,
            "inst_salud":    "",
            "renta":         renta,
            "cot_afp":       cot_afp,
            "sis":           sis,
            "afc":           0.0,
            "cot_salud":     0.0,
            "mutual":        0.0,
            "ccaf":          0.0,
            "dias":          0,
            "tipo_contrato": tipo,
            "cargas_familiar": 0.0,
        })
    return registros


# ── IPS Anexo Trabajadores ───────────────────────────────────────────────────
# Estructura (20 columnas):
# [0]N° | [1]Rut(sin DV) | [2]DV | [3]Apellido Paterno, Materno, Nombres |
# [4]Dias | [5]Rem.Imponible | [6]Pensiones IPS | [7]Fonasa | [8]Accidentes |
# [9]Rem.Desahucio | [10]Cot.Desahucio | [11]Cod.Mov | [12]F.Inicio |
# [13]F.Término | [14]Tramo | [15]Simple(N°) | [16]Inválida(N°) |
# [17]Maternal(N°) | [18]Monto Asig.Familiar | [19]Bonif.
#
# Aporta: cargas_familiares (monto col 18) + nombre como respaldo
# ────────────────────────────────────────────────────────────────────────────

def parsear_ips(tabla: list) -> list[dict]:
    registros = []
    for fila in tabla:
        if not fila or len(fila) < 15:
            continue
        # RUT dividido: [1]=número sin DV, [2]=DV
        rut_num = str(fila[1]).strip() if fila[1] else ""
        dv      = str(fila[2]).strip() if fila[2] else ""
        if not rut_num or not rut_num.replace(".", "").isdigit():
            continue
        rut_raw = f"{rut_num}-{dv}"
        if not es_rut_valido(rut_raw):
            continue

        nombre_raw      = str(fila[3]).strip() if fila[3] else ""
        cargas_familiar = a_numero(fila[18]) if len(fila) > 18 else 0.0

        # Ignorar filas de totales
        if "total" in nombre_raw.lower():
            continue

        registros.append({
            "fuente":          "ips_anexo",
            "rut":             limpiar_rut(rut_raw),
            "nombre":          nombre_raw,
            "inst_afp":        "",
            "inst_salud":      "",
            "renta":           0.0,
            "cot_afp":         0.0,
            "sis":             0.0,
            "afc":             0.0,
            "cot_salud":       0.0,
            "seg_social":      0.0,
            "mutual":          0.0,
            "ccaf":            0.0,
            "dias":            0,
            "tipo_contrato":   "",
            "cargas_familiar": cargas_familiar,
        })
    return registros
# Columnas: RUT | Ap.Paterno | Ap.Materno | Nombres | Remuneración | Cotización
#
# Esta fuente aporta: mutual + nombre (si no vino del AFP)
# La renta aquí se IGNORA (ya viene del AFP para no duplicar)
# ────────────────────────────────────────────────────────────────────────────

def parsear_mutual(tabla: list) -> list[dict]:
    registros = []
    for fila in tabla:
        if not fila or es_encabezado(fila):
            continue
        rut_raw, off = _rut_offset(fila)
        if not rut_raw:
            continue

        nom = nombre_completo(
            fila[off + 1] if len(fila) > off + 1 else "",
            fila[off + 2] if len(fila) > off + 2 else "",
            fila[off + 3] if len(fila) > off + 3 else "",
        )
        mutual = a_numero(fila[off + 5]) if len(fila) > off + 5 else 0.0

        if mutual == 0:
            continue

        registros.append({
            "fuente":        "mutual",
            "rut":           limpiar_rut(rut_raw),
            "nombre":        nom,
            "inst_afp":      "",
            "inst_salud":    "Mutual CChC",
            "renta":         0.0,
            "cot_afp":       0.0,
            "sis":           0.0,
            "afc":           0.0,
            "cot_salud":     0.0,
            "mutual":        mutual,
            "ccaf":          0.0,
            "dias":          0,
            "tipo_contrato": "",
            "cargas_familiar": 0.0,
        })
    return registros


# ── Fonasa Anexo ─────────────────────────────────────────────────────────────
# Columnas: N° | RUT | Ap.Paterno | Materno | Nombres | Días | Entidad |
#           Rem.Imponible | Cotización 7% | Cod.Mov | ...
#
# Esta fuente aporta: cot_salud + nombre (si no vino del AFP)
# La renta aquí se IGNORA (ya viene del AFP)
# ────────────────────────────────────────────────────────────────────────────

def parsear_fonasa(tabla: list) -> list[dict]:
    """
    Tabla Fonasa (ANEXOS DE DETALLE):
    N° | RUT | Ap.Paterno | Materno | Nombres | Días | Entidad |
    Rem.Imponible | Cotización 7% | Cod.Mov | ...

    Aporta: cot_salud + días trabajados.
    Tipo contrato se detecta en parsear_afp (cotización afiliado AFC > 0).
    """
    registros = []
    for fila in tabla:
        if not fila or es_encabezado(fila):
            continue
        rut_raw, off = _rut_offset(fila)
        if not rut_raw:
            continue

        nom = nombre_completo(
            fila[off + 1] if len(fila) > off + 1 else "",
            fila[off + 2] if len(fila) > off + 2 else "",
            fila[off + 3] if len(fila) > off + 3 else "",
        )
        dias    = int(a_numero(fila[off + 4])) if len(fila) > off + 4 else 0
        cot_sal = a_numero(fila[off + 7]) if len(fila) > off + 7 else 0.0

        if cot_sal == 0:
            continue

        registros.append({
            "fuente":        "fonasa",
            "rut":           limpiar_rut(rut_raw),
            "nombre":        nom,
            "inst_afp":      "",
            "inst_salud":    "Fonasa",
            "renta":         0.0,
            "cot_afp":       0.0,
            "sis":           0.0,
            "afc":           0.0,
            "cot_salud":     cot_sal,
            "mutual":        0.0,
            "ccaf":          0.0,
            "dias":          dias,
            "tipo_contrato": "",   # viene del AFP
            "cargas_familiar": 0.0,
        })
    return registros


# ── Seguro Social ────────────────────────────────────────────────────────────
# Estructura real (13 columnas):
# [0]N° | [1]RUT | [2]Nombres y Apellidos | [3]Renta Imponible |
# [4]Renta Imp.Ant./Días Lic. | [5]Seg.Social Previsional | [6]Rent.Protegida |
# [7]SIS | [8]Tipo Jornada | [9]Días Trabajados | [10]Código Mov. |
# [11]Fecha Inicio | [12]Fecha Término
#
# Aporta: seg_social (cotización propia) + nombre como respaldo
# Renta y días se IGNORAN (ya vienen de AFP y Fonasa respectivamente)
# ────────────────────────────────────────────────────────────────────────────

def parsear_seguro_social(tabla: list) -> list[dict]:
    registros = []
    for fila in tabla:
        if not fila or len(fila) < 6:
            continue
        # Col [1] es siempre el RUT (col [0] es N° de fila)
        rut_raw = str(fila[1]).strip() if fila[1] else ""
        if not es_rut_valido(rut_raw):
            continue

        nombre_raw  = str(fila[2]).strip() if fila[2] else ""
        seg_social  = a_numero(fila[5]) if len(fila) > 5 else 0.0

        # Ignorar filas de totales (nombre = "Totales Generales")
        if "total" in nombre_raw.lower():
            continue

        registros.append({
            "fuente":        "seguro_social",
            "rut":           limpiar_rut(rut_raw),
            "nombre":        nombre_raw,
            "inst_afp":      "",
            "inst_salud":    "",
            "renta":         0.0,      # viene del AFP
            "cot_afp":       0.0,
            "sis":           0.0,      # viene del AFP
            "afc":           0.0,
            "cot_salud":     0.0,      # viene de Fonasa
            "seg_social":    seg_social,  # campo exclusivo de esta fuente
            "mutual":        0.0,
            "ccaf":          0.0,
            "dias":          0,        # viene de Fonasa
            "tipo_contrato": "",
        })
    return registros


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 4: EXTRACCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def extraer_datos_pdf(ruta_pdf: str) -> list[dict]:
    """
    Recorre cada página del PDF, detecta su tipo y llama al parser correspondiente.
    Agrega 'pdf_origen' a cada registro para poder deduplicar días por PDF.
    """
    registros = []
    pdf_origen = os.path.basename(ruta_pdf)   # nombre del archivo, identifica la obra

    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text() or ""
                tipo  = detectar_tipo_pagina(texto)
                if tipo == "otro":
                    continue

                tablas = pagina.extract_tables()
                if not tablas:
                    continue

                afp_nombre = detectar_afp_en_texto(texto)

                for tabla in tablas:
                    if not tabla:
                        continue
                    nuevos = []
                    if tipo == "afp_detalle":
                        nuevos = parsear_afp(tabla, afp_nombre)
                    elif tipo == "ips_anexo":
                        nuevos = parsear_ips(tabla)
                    elif tipo == "mutual":
                        nuevos = parsear_mutual(tabla)
                    elif tipo == "fonasa_anexo":
                        nuevos = parsear_fonasa(tabla)
                    elif tipo == "seguro_social":
                        nuevos = parsear_seguro_social(tabla)

                    # Marcar cada registro con el PDF que lo originó
                    for r in nuevos:
                        r["pdf_origen"] = pdf_origen
                    registros.extend(nuevos)

    except Exception as e:
        print(f"  [ERROR] {ruta_pdf}: {e}")

    return registros


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 5: CONSOLIDACIÓN CON PANDAS
# ─────────────────────────────────────────────────────────────────────────────
#
# REGLA CLAVE:
#   - "renta" solo la suman los registros de fuente "afp" (los demás tienen 0)
#   - "cot_afp" + "afc"   → solo fuente "afp"
#   - "cot_salud"          → solo fuente "fonasa"
#   - "mutual"             → solo fuente "mutual"
#   - Los registros de seguro_social y otros tienen todos los montos en 0
#     (solo aportan el nombre por si no vino de otra fuente)
#
#   Al hacer groupby().sum() esto funciona automáticamente porque
#   cada campo solo tiene valor en la fuente que le corresponde.
#
# ─────────────────────────────────────────────────────────────────────────────

COLS_NUM  = ["renta", "cot_afp", "sis", "cot_salud", "seg_social", "afc", "mutual", "ccaf", "cargas_familiar"]
COLS_TEXT = ["inst_afp", "inst_salud"]


def primer_no_vacio(serie):
    vals = serie.dropna().astype(str)
    vals = vals[vals.str.strip().str.len() > 0]
    return vals.iloc[0] if not vals.empty else ""


def nombre_mas_largo(serie):
    vals = serie.dropna().astype(str)
    vals = vals[vals.str.strip().str.len() > 0]
    return max(vals, key=len) if not vals.empty else ""


def tipo_contrato_predominante(serie):
    """
    Indefinido solo si el RUT está en RUT_INDEFINIDO.
    En cualquier otro caso → Plazo Fijo.
    """
    vals = serie.dropna().astype(str)
    vals = vals[vals.isin(["Indefinido", "Plazo Fijo"])]
    if vals.empty:
        return "Plazo Fijo"
    if (vals == "Indefinido").any():
        return "Indefinido"
    return "Plazo Fijo"


def consolidar_datos(registros: list[dict]) -> pd.DataFrame:
    """
    Consolidación en DOS pasos para cruzar correctamente la información:

    PASO 1 — Colapsar por (rut + pdf_origen):
        Dentro de un mismo PDF el trabajador aparece en AFP, Fonasa, Mutual y
        Seguro Social. Cada fuente aporta su campo exclusivo (renta→AFP,
        cot_salud→Fonasa, mutual→Mutual). Los días vienen de Fonasa (una vez
        por PDF). Al agrupar por rut+pdf_origen obtenemos una fila por
        trabajador por PDF, con todos sus campos correctos y sin duplicados.

    PASO 2 — Colapsar por rut:
        Un trabajador puede aparecer en varios PDFs (distintas obras). Aquí sí
        se suman renta, cotizaciones Y días, porque son obras diferentes.
    """
    if not registros:
        return pd.DataFrame()

    df = pd.DataFrame(registros)

    # Asegurar columnas numéricas
    for col in COLS_NUM + ["dias"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    for col in COLS_TEXT:
        if col not in df.columns:
            df[col] = ""

    if "pdf_origen" not in df.columns:
        df["pdf_origen"] = "desconocido"
    if "tipo_contrato" not in df.columns:
        df["tipo_contrato"] = ""

    # ── PASO 1: un registro por (rut, pdf_origen) ────────────────────────────
    agg_paso1 = {
        "nombre":         nombre_mas_largo,
        **{c: "sum"            for c in COLS_NUM},
        **{c: primer_no_vacio  for c in COLS_TEXT},
        "dias":           "sum",
        "tipo_contrato":  tipo_contrato_predominante,
    }
    df_por_pdf = (
        df.groupby(["rut", "pdf_origen"], as_index=False)
          .agg(agg_paso1)
    )

    # ── PASO 2: un registro por rut (suma entre obras/PDFs distintos) ─────────
    agg_paso2 = {
        "nombre":         nombre_mas_largo,
        **{c: "sum"            for c in COLS_NUM},
        **{c: primer_no_vacio  for c in COLS_TEXT},
        "dias":           "sum",
        "tipo_contrato":  tipo_contrato_predominante,
    }
    df_out = (
        df_por_pdf.groupby("rut", as_index=False)
                  .agg(agg_paso2)
                  .sort_values("nombre", ignore_index=True)
    )

    # ── AFC según tipo de contrato ────────────────────────────────────────────
    # Indefinido:  trabajador 0.6% + empleador 2.4% = 3.0% total
    # Plazo Fijo:  solo empleador 3.0% (trabajador no cotiza AFC)
    # Tasa total declarada = 3% en ambos casos → cálculo idéntico sobre renta
    df_out["afc"] = (df_out["renta"] * 0.03).round(0)

    df_out.rename(columns={
        "rut":             "RUT",
        "nombre":          "Nombre Completo",
        "inst_salud":      "Institución Salud",
        "inst_afp":        "Institución AFP",
        "tipo_contrato":   "Tipo Contrato",
        "renta":           "Renta Imponible",
        "cot_afp":         "Cotización AFP",
        "sis":             "SIS",
        "cot_salud":       "Cotización Salud (Fonasa)",
        "seg_social":      "Seguro Social",
        "afc":             "Seguro Cesantía (AFC)",
        "mutual":          "Mutual",
        "ccaf":            "CCAF",
        "cargas_familiar": "Cargas Familiares",
        "dias":            "Días Trabajados",
    }, inplace=True)

    cols_orden = [
        "RUT", "Nombre Completo", "Institución AFP", "Institución Salud",
        "Tipo Contrato", "Días Trabajados",
        "Renta Imponible", "Cotización AFP", "SIS",
        "Cotización Salud (Fonasa)", "Seguro Social",
        "Seguro Cesantía (AFC)", "Mutual", "CCAF", "Cargas Familiares",
    ]
    df_out = df_out[[c for c in cols_orden if c in df_out.columns]]

    return df_out


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 6: EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────────────────────────────────────

COLS_DINERO = {
    "Renta Imponible", "Cotización AFP", "SIS",
    "Cotización Salud (Fonasa)", "Seguro Social",
    "Seguro Cesantía (AFC)", "Mutual", "CCAF", "Cargas Familiares"
}


def exportar_excel(df: pd.DataFrame, ruta: str) -> None:
    df.to_excel(ruta, index=False, sheet_name="Consolidado Previred")
    wb = load_workbook(ruta)
    ws = wb.active

    COLOR_HDR = "1F3864"
    COLOR_PAR = "DCE6F1"
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for cel in ws[1]:
        cel.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cel.fill      = PatternFill("solid", fgColor=COLOR_HDR)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border    = borde
    ws.row_dimensions[1].height = 30

    encabezados = [c.value for c in ws[1]]

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2), start=2):
        color = COLOR_PAR if num_fila % 2 == 0 else "FFFFFF"
        for cel in fila:
            cel.font   = Font(name="Arial", size=9)
            cel.fill   = PatternFill("solid", fgColor=color)
            cel.border = borde
            col_nom    = encabezados[cel.column - 1] if cel.column <= len(encabezados) else ""
            if col_nom in COLS_DINERO:
                cel.number_format = "#,##0"
                cel.alignment = Alignment(horizontal="right", vertical="center")
            elif col_nom in ("Días Trabajados", "Tipo Contrato"):
                cel.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cel.alignment = Alignment(horizontal="left", vertical="center")

    for i, cols in enumerate(ws.columns, 1):
        ancho = max((len(str(c.value or "")) for c in cols), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(ancho + 3, 12), 40)

    ws.freeze_panes = "A2"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen Consolidado Previred"
    ws2["A1"].font = Font(bold=True, size=12)
    filas_res = [
        ("Total trabajadores únicos",      len(df)),
        ("Total Renta Imponible",           df["Renta Imponible"].sum()),
        ("Total Cotización AFP",            df["Cotización AFP"].sum()),
        ("Total SIS",                       df["SIS"].sum()),
        ("Total Cotización Salud (Fonasa)", df["Cotización Salud (Fonasa)"].sum()),
        ("Total Seguro Social",             df["Seguro Social"].sum()),
        ("Total Seg. Cesantía AFC (3%)",    df["Seguro Cesantía (AFC)"].sum()),
        ("Total Mutual",                    df["Mutual"].sum()),
        ("Total Cargas Familiares",         df["Cargas Familiares"].sum()),
        ("Contratos Indefinidos",           (df["Tipo Contrato"] == "Indefinido").sum()),
        ("Contratos Plazo Fijo",            (df["Tipo Contrato"] == "Plazo Fijo").sum()),
        ("Promedio Días Trabajados",        round(df["Días Trabajados"].mean(), 1)),
    ]
    for i, (label, val) in enumerate(filas_res, start=3):
        ws2.cell(i, 1, label)
        c = ws2.cell(i, 2, val)
        if i > 3:
            c.number_format = "#,##0"
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    wb.save(ruta)


# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 7: INTERFAZ GRÁFICA
# ─────────────────────────────────────────────────────────────────────────────


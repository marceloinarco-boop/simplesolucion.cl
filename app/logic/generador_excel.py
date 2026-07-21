"""
====================================================
  GENERADOR EXCEL - LIQUIDACIÃN DE SUELDO
====================================================
Genera una liquidaciÃ³n de sueldo en formato .xlsx
con formato profesional usando openpyxl.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from .constantes import _get_directorio_exportacion

MESES_ES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

# ââ Paleta de colores (hex sin #) ââââââââââââââââââââââââââââââââ
C_PRIMARIO   = '0f2942'   # azul marino
C_SECUNDARIO = '0ea5e9'   # azul claro
C_EXITO      = '16a34a'   # verde
C_ERROR      = 'dc2626'   # rojo
C_ACENTO     = 'e8f4f8'   # celeste muy suave
C_GRIS_HDR   = 'f1f5f9'   # fondo cabecera
C_BORDE      = 'd1d5db'
BLANCO       = 'FFFFFF'


def _thin_border(left=True, right=True, top=True, bottom=True):
    s = Side(style='thin', color=C_BORDE)
    n = Side(style=None)
    return Border(
        left=s if left else n,
        right=s if right else n,
        top=s if top else n,
        bottom=s if bottom else n,
    )


def _thick_bottom():
    return Border(bottom=Side(style='medium', color=C_PRIMARIO))


def _fmt_pesos(v) -> str:
    try:
        return f"$ {int(round(float(v))):,}".replace(',', '.')
    except Exception:
        return "$ 0"


def _nombre_mes(num: int) -> str:
    return MESES_ES[num] if 1 <= num <= 12 else str(num)


# ââ Estilos reutilizables ââââââââââââââââââââââââââââââââââââââââ
def _f(bold=False, size=10, color='000000', name='Arial'):
    return Font(name=name, bold=bold, size=size, color=color)


def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)


def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _left():
    return Alignment(horizontal='left', vertical='center', indent=1)


def _right():
    return Alignment(horizontal='right', vertical='center')


# ââ Helpers de escritura âââââââââââââââââââââââââââââââââââââââââ
def _write(ws, row, col, value, font=None, fill=None,
           alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:         cell.font = font
    if fill:         cell.fill = fill
    if alignment:    cell.alignment = alignment
    if border:       cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def _merge_write(ws, row, col_start, col_end, value,
                 font=None, fill=None, alignment=None, border=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border = border
    return cell


# ââ Generador principal ââââââââââââââââââââââââââââââââââââââââââ
class GeneradorExcel:

    TOTAL_COLS = 6   # A:F â etiqueta | detalle | importe | | etiqueta2 | importe2

    def __init__(self, trabajador: dict, entrada: dict, resultado: dict,
                 empresa: dict = None, periodo: str = None, ajustes: dict = None):
        self.trabajador = trabajador
        self.entrada    = entrada
        self.res        = resultado
        self.empresa    = empresa or {
            'nombre': 'Estructuras Alejandra Fortuzzi',
            'rut': '76.000.000-0',
            'direccion': '',
            'ciudad': 'Santiago',
        }
        self.periodo = periodo or datetime.now().strftime('%Y-%m')
        self.ajustes = ajustes or {}

        try:
            anio, mes = self.periodo.split('-')
            self.anio = int(anio)
            self.mes  = int(mes)
        except Exception:
            self.anio = datetime.now().year
            self.mes  = datetime.now().month

    # ââââââââââââââââââââââââââââââââââââââââââââââââ
    def generar(self) -> str:
        directorio = _get_directorio_exportacion(self.anio, self.mes)
        os.makedirs(directorio, exist_ok=True)

        nombre_archivo = (
            f"Liquidacion_{self.trabajador['nombre'].replace(' ', '_')}"
            f"_{self.periodo}.xlsx"
        )
        ruta = os.path.join(directorio, nombre_archivo)

        wb = Workbook()
        ws = wb.active
        ws.title = f"LiquidaciÃ³n {self.periodo}"

        self._configurar_columnas(ws)
        row = self._encabezado(ws, 1)
        row = self._datos_generales(ws, row)
        row = self._haberes(ws, row)
        row = self._descuentos(ws, row)
        row = self._liquido(ws, row)
        row = self._base_imponible(ws, row)
        self._pie(ws, row)

        # Congelar fila 1
        ws.freeze_panes = 'A5'
        ws.sheet_view.showGridLines = False

        wb.save(ruta)
        return ruta

    # ââ Columnas âââââââââââââââââââââââââââââââââââââ
    def _configurar_columnas(self, ws):
        widths = [28, 18, 16, 2, 28, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ââ Banda de tÃ­tulo empresa âââââââââââââââââââââââ
    def _encabezado(self, ws, row: int) -> int:
        # Fila 1: logo/empresa (fondo marino)
        ws.row_dimensions[row].height = 32
        _merge_write(ws, row, 1, 6,
                     self.empresa.get('nombre', ''),
                     font=_f(bold=True, size=14, color=BLANCO, name='Arial'),
                     fill=_fill(C_PRIMARIO),
                     alignment=_center())
        row += 1

        # Fila 2: RUT empresa + direcciÃ³n
        ws.row_dimensions[row].height = 18
        rut_emp = self.empresa.get('rut', '')
        dir_emp = self.empresa.get('direccion', '')
        ciudad  = self.empresa.get('ciudad', 'Santiago')
        subtitulo = f"RUT: {rut_emp}   |   {dir_emp}, {ciudad}"
        _merge_write(ws, row, 1, 6, subtitulo,
                     font=_f(size=9, color='d0e8f5', name='Arial'),
                     fill=_fill(C_PRIMARIO),
                     alignment=_center())
        row += 1

        # Fila 3: TÃ­tulo liquidaciÃ³n + perÃ­odo
        ws.row_dimensions[row].height = 24
        periodo_txt = f"{_nombre_mes(self.mes)} {self.anio}"
        _merge_write(ws, row, 1, 4,
                     "LIQUIDACIÃN DE SUELDO",
                     font=_f(bold=True, size=13, color=C_PRIMARIO, name='Arial'),
                     fill=_fill(C_ACENTO),
                     alignment=Alignment(horizontal='left', vertical='center', indent=1))
        _merge_write(ws, row, 5, 6,
                     f"PerÃ­odo: {periodo_txt}",
                     font=_f(bold=True, size=11, color=C_SECUNDARIO, name='Arial'),
                     fill=_fill(C_ACENTO),
                     alignment=_right())
        # LÃ­nea gruesa debajo del tÃ­tulo
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _thick_bottom()
        row += 1
        return row

    # ââ Datos generales (empleado + empresa) âââââââââ
    def _datos_generales(self, ws, row: int) -> int:
        ws.row_dimensions[row].height = 18
        _merge_write(ws, row, 1, 6, "DATOS DEL TRABAJADOR",
                     font=_f(bold=True, size=9, color=BLANCO),
                     fill=_fill(C_SECUNDARIO),
                     alignment=_left())
        row += 1

        t = self.trabajador
        e = self.entrada
        filas_gen = [
            ("Nombre", t.get('nombre', ''),
             "Tipo de Contrato", e.get('tipo_contrato', '')),
            ("RUT Trabajador", t.get('rut', ''),
             "AFP", e.get('afp', '')),
            ("Salud", e.get('tipo_salud', 'FONASA'),
             "DÃ­as Trabajados", str(e.get('dias_trabajados', 30))),
        ]
        obra = e.get('nombre_obra', '').strip()
        if obra:
            filas_gen.append(("Obra", obra, "", ""))
        for lbl1, val1, lbl2, val2 in filas_gen:
            ws.row_dimensions[row].height = 17
            _write(ws, row, 1, lbl1,
                   font=_f(bold=True, size=9, color='374151'),
                   fill=_fill(C_GRIS_HDR),
                   alignment=_left(),
                   border=_thin_border())
            _write(ws, row, 2, val1,
                   font=_f(size=9),
                   fill=_fill(BLANCO),
                   alignment=_left(),
                   border=_thin_border())
            # Espacio visual
            ws.cell(row=row, column=3).fill = _fill(BLANCO)
            ws.cell(row=row, column=4).fill = _fill(BLANCO)
            _write(ws, row, 5, lbl2,
                   font=_f(bold=True, size=9, color='374151'),
                   fill=_fill(C_GRIS_HDR),
                   alignment=_left(),
                   border=_thin_border())
            _write(ws, row, 6, val2,
                   font=_f(size=9),
                   fill=_fill(BLANCO),
                   alignment=_left(),
                   border=_thin_border())
            row += 1

        # Espacio
        ws.row_dimensions[row].height = 8
        row += 1
        return row

    # ââ SecciÃ³n helper ââââââââââââââââââââââââââââââââ
    def _titulo_seccion(self, ws, row, titulo, color_bg=None):
        color_bg = color_bg or C_PRIMARIO
        ws.row_dimensions[row].height = 18
        _merge_write(ws, row, 1, 6, titulo,
                     font=_f(bold=True, size=9, color=BLANCO),
                     fill=_fill(color_bg),
                     alignment=_left())
        return row + 1

    def _fila_concepto(self, ws, row, concepto, detalle='', monto=None,
                       bold=False, color_monto='000000', bg=BLANCO):
        ws.row_dimensions[row].height = 16
        _write(ws, row, 1, concepto,
               font=_f(bold=bold, size=9, color='1e293b'),
               fill=_fill(bg),
               alignment=_left(),
               border=_thin_border(right=False, top=False))
        _write(ws, row, 2, detalle,
               font=_f(size=8, color='64748b'),
               fill=_fill(bg),
               alignment=_left(),
               border=_thin_border(left=False, right=False, top=False))
        # Celda monto (cols 3 y 4 fusionadas)
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=4)
        cell_m = ws.cell(row=row, column=3)
        if monto is not None:
            cell_m.value = monto
            cell_m.number_format = '"$ "#,##0'
        cell_m.font     = _f(bold=bold, size=9, color=color_monto)
        cell_m.fill     = _fill(bg)
        cell_m.alignment = _right()
        cell_m.border   = _thin_border(left=False, top=False)
        # Cols 5-6 vacÃ­as pero formateadas
        for c in [5, 6]:
            ws.cell(row=row, column=c).fill = _fill(bg)
        return row + 1

    def _fila_total(self, ws, row, etiqueta, formula_or_val, color_fondo, color_txt):
        ws.row_dimensions[row].height = 20
        _merge_write(ws, row, 1, 2, etiqueta,
                     font=_f(bold=True, size=10, color=color_txt),
                     fill=_fill(color_fondo),
                     alignment=_left())
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=4)
        cell = ws.cell(row=row, column=3, value=formula_or_val)
        cell.font        = _f(bold=True, size=11, color=color_txt)
        cell.fill        = _fill(color_fondo)
        cell.alignment   = _right()
        cell.number_format = '"$ "#,##0'
        for c in [5, 6]:
            ws.cell(row=row, column=c).fill = _fill(color_fondo)
        return row + 1

    # ââ Haberes âââââââââââââââââââââââââââââââââââââââ
    def _haberes(self, ws, row: int) -> int:
        row = self._titulo_seccion(ws, row, "HABERES")
        r = self.res
        e = self.entrada

        hab_start = row   # para formula SUM

        row = self._fila_concepto(ws, row, "Sueldo Base",
                                   monto=float(r.get('sueldo_base', e.get('sueldo_base', 0))))
        grat = r.get('gratificacion', 0)
        if grat:
            row = self._fila_concepto(ws, row, "GratificaciÃ³n legal (25%)",
                                       monto=float(grat))
        hex_m = r.get('horas_extra_monto', 0)
        if hex_m:
            row = self._fila_concepto(ws, row,
                                       f"Horas Extra ({e.get('horas_extra', 0)} hrs)",
                                       monto=float(hex_m))
        bono = e.get('bono_imponible', 0)
        if bono:
            row = self._fila_concepto(ws, row, "Bonos Imponibles",
                                       monto=float(bono))
        colac = e.get('colacion', 0)
        if colac:
            row = self._fila_concepto(ws, row, "ColaciÃ³n (no imponible)",
                                       monto=float(colac))
        movil = e.get('movilizacion', 0)
        if movil:
            row = self._fila_concepto(ws, row, "MovilizaciÃ³n (no imponible)",
                                       monto=float(movil))
        viat = e.get('viaticos', 0)
        if viat:
            row = self._fila_concepto(ws, row, "ViÃ¡ticos",
                                       monto=float(viat))
        asig = r.get('asig_familiar', 0)
        if asig:
            nc = e.get('cargas_familiares', 0)
            row = self._fila_concepto(ws, row,
                                       f"Asig. Familiar ({nc} carga(s))",
                                       monto=float(asig))

        hab_end = row - 1
        total_hab = float(r.get('total_haberes', 0))
        row = self._fila_total(ws, row, "TOTAL HABERES", total_hab,
                                'dbeafe', C_SECUNDARIO)
        ws.row_dimensions[row - 1].height = 20

        ws.row_dimensions[row].height = 8
        row += 1
        return row

    # ââ Descuentos ââââââââââââââââââââââââââââââââââââ
    def _descuentos(self, ws, row: int) -> int:
        row = self._titulo_seccion(ws, row, "DESCUENTOS PREVISIONALES", 'b91c1c')
        r = self.res
        e = self.entrada

        tasa_afp = r.get('tasa_afp', 0)
        row = self._fila_concepto(ws, row,
                                   f"AFP  {e.get('afp', '')}",
                                   f"{tasa_afp*100:.2f}%".replace('.', ','),
                                   monto=float(r.get('afp_monto', 0)),
                                   color_monto=C_ERROR)
        salud_lbl = e.get('tipo_salud', 'FONASA')
        row = self._fila_concepto(ws, row,
                                   f"Salud  {salud_lbl}",
                                   "7.00%",
                                   monto=float(r.get('salud_monto', 0)),
                                   color_monto=C_ERROR)
        ces = r.get('cesantia_trabajador', 0)
        if ces:
            row = self._fila_concepto(ws, row, "Seguro de CesantÃ­a (trabajador)",
                                       monto=float(ces), color_monto=C_ERROR)
        imp = r.get('impuesto_segunda_cat', 0)
        if imp:
            row = self._fila_concepto(ws, row, "Impuesto 2Âª CategorÃ­a",
                                       monto=float(imp), color_monto=C_ERROR)

        total_desc = float(r.get('total_descuentos', 0))
        row = self._fila_total(ws, row, "TOTAL DESCUENTOS", total_desc,
                                'fee2e2', C_ERROR)

        ws.row_dimensions[row].height = 8
        row += 1
        return row

    # ââ Sueldo LÃ­quido ââââââââââââââââââââââââââââââââ
    def _liquido(self, ws, row: int) -> int:
        ws.row_dimensions[row].height = 28
        _merge_write(ws, row, 1, 2,
                     "SUELDO LÃQUIDO A PAGAR",
                     font=_f(bold=True, size=12, color=BLANCO),
                     fill=_fill(C_EXITO),
                     alignment=_left())
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=4)
        cell = ws.cell(row=row, column=3,
                       value=float(self.res.get('sueldo_liquido', 0)))
        cell.font        = _f(bold=True, size=14, color=BLANCO)
        cell.fill        = _fill(C_EXITO)
        cell.alignment   = _right()
        cell.number_format = '"$ "#,##0'
        for c in [5, 6]:
            ws.cell(row=row, column=c).fill = _fill(C_EXITO)

        ws.row_dimensions[row + 1].height = 8
        return row + 2

    # ââ Base imponible ââââââââââââââââââââââââââââââââ
    def _base_imponible(self, ws, row: int) -> int:
        row = self._titulo_seccion(ws, row, "BASE IMPONIBLE", '374151')
        r = self.res
        row = self._fila_concepto(ws, row, "Base Previsional (AFP / Salud)",
                                   monto=float(r.get('base_previsional', 0)),
                                   bg=C_GRIS_HDR)
        row = self._fila_concepto(ws, row, "Base Tributable (Imp. 2Âª Cat.)",
                                   monto=float(r.get('base_tributable', 0)),
                                   bg=C_GRIS_HDR)
        ws.row_dimensions[row].height = 8
        return row + 1

    # ââ Pie âââââââââââââââââââââââââââââââââââââââââââ
    def _pie(self, ws, row: int):
        ws.row_dimensions[row].height = 16
        pie_txt = self.ajustes.get('pdf_pie_texto', 'Estructuras Alejandra Fortuzzi')
        _merge_write(ws, row, 1, 4, pie_txt,
                     font=_f(size=8, color='94a3b8'),
                     fill=_fill(BLANCO),
                     alignment=_left())
        _merge_write(ws, row, 5, 6, "by informagic",
                     font=_f(size=8, color='94a3b8'),
                     fill=_fill(BLANCO),
                     alignment=_right())


# ââ FunciÃ³n pÃºblica âââââââââââââââââââââââââââââââââââââââââââââââ
def generar_excel_liquidacion(trabajador: dict, entrada: dict,
                               resultado: dict, empresa: dict = None,
                               periodo: str = None,
                               ajustes: dict = None) -> str:
    gen = GeneradorExcel(trabajador, entrada, resultado,
                         empresa, periodo, ajustes)
    return gen.generar()
